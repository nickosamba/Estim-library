from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from .models import Book, Category, Favorite, ReadingProgress, Review, Bookmark, Annotation, Chapter, Campus
from .forms import ReviewForm, BookForm, AuthorForm
from django.db.models import Q, Avg
from django.contrib.auth import login, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
import requests
import re
import csv
import io
import openpyxl

User = get_user_model()

from django.views.decorators.http import require_GET
from django.conf import settings
import os

@require_GET
def service_worker(request):
    sw_path = os.path.join(settings.BASE_DIR, 'static', 'sw.js')
    with open(sw_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/javascript')
        return response

def book_list(request):
    """
    Affiche la page d'accueil avec les livres, recommandations et statistiques.
    Optimisé pour charger les relations auteur et catégorie en une seule requête.
    """
    query = request.GET.get('q')
    if query:
        # Recherche plein texte simple sur le titre, l'auteur et la catégorie
        books = Book.objects.filter(
            Q(title__icontains=query) | 
            Q(author__name__icontains=query) |
            Q(category__name__icontains=query),
            is_available=True
        ).select_related('author', 'category')
    else:
        books = Book.objects.filter(is_available=True).select_related('author', 'category')

    # Dernières nouveautés (6 derniers livres ajoutés)
    latest_books = Book.objects.filter(is_available=True).select_related('author', 'category').order_by('-created_at')[:6]

    # Statistiques réelles pour les compteurs de la page d'accueil
    total_books = Book.objects.count()
    total_members = User.objects.count()

    # Recommandations intelligentes basées sur le profil (Campus + Département)
    recommendations = Book.objects.none()
    if request.user.is_authenticated:
        # Priorité aux livres du campus de l'utilisateur ou "Tout Public"
        recommendations = Book.objects.filter(
            Q(target_campuses=request.user.campus) | Q(target_campuses__code='all'),
            is_available=True
        ).select_related('author', 'category')
        
        # Filtre optionnel par département si renseigné dans le profil
        if request.user.department:
            recommendations = recommendations.filter(target_department=request.user.department)
        
        # Sélection aléatoire parmi les correspondances
        recommendations = recommendations.distinct().order_by('?')[:6]
    else:
        # Fallback pour anonymes : sélection aléatoire d'ouvrages disponibles
        recommendations = Book.objects.filter(is_available=True).select_related('author', 'category').order_by('?')[:6]
    
    # Logique du "Trésor du Mois" (Hiérarchie de sélection)
    # 1. Livres marqués explicitement "Featured"
    treasure_of_the_month = Book.objects.filter(is_featured=True, is_available=True).exclude(cover_image='').select_related('author', 'category').order_by('-updated_at').first()
    
    # 2. Fallback : Les mieux notés (>= 4 étoiles)
    if not treasure_of_the_month:
        treasure_of_the_month = Book.objects.annotate(
            avg_rating=Avg('reviews__rating')
        ).filter(avg_rating__gte=4, is_available=True).exclude(cover_image='').select_related('author', 'category').order_by('?').first()
    
    # 3. Dernier recours : N'importe quel livre avec une couverture
    if not treasure_of_the_month:
        treasure_of_the_month = Book.objects.filter(is_available=True).exclude(cover_image='').select_related('author', 'category').order_by('?').first()

    # Aperçu de la progression de lecture (3 derniers ouvrages consultés)
    reading_progress = []
    if request.user.is_authenticated:
        reading_progress = ReadingProgress.objects.filter(user=request.user).select_related('book').order_by('-updated_at')[:3]
    
    context = {
        'books': books,
        'latest_books': latest_books,
        'recommendations': recommendations,
        'treasure_of_the_month': treasure_of_the_month,
        'reading_progress': reading_progress,
        'search_query': query,
        'total_books': total_books,
        'total_members': total_members,
    }
    return render(request, 'books/book_list.html', context)

def book_detail(request, slug):
    """
    Affiche la fiche détaillée d'un ouvrage.
    Gère les avis, les favoris et la progression de lecture.
    """
    book = get_object_or_404(Book.objects.select_related('author', 'category'), slug=slug)
    
    is_favorite = False
    if request.user.is_authenticated:
        is_favorite = Favorite.objects.filter(user=request.user, book=book).exists()
    
    # Récupération des avis (optimisé avec l'utilisateur)
    reviews = book.reviews.all().select_related('user')
    user_review = None
    if request.user.is_authenticated:
        user_review = reviews.filter(user=request.user).first()
    
    form = ReviewForm(instance=user_review)
    
    # Vérification si le livre est disponible physiquement sur le campus de l'utilisateur
    is_local = True
    if request.user.is_authenticated:
        is_local = book.is_available_at(request.user.campus)
    
    # Recommandations contextuelles (même catégorie, excluant le livre actuel)
    similar_books = Book.objects.filter(category=book.category).exclude(id=book.id).select_related('author')[:4]
    
    # Progression de lecture spécifique à l'ouvrage
    reading_progress = None
    if request.user.is_authenticated:
        reading_progress = ReadingProgress.objects.filter(user=request.user, book=book).first()

    context = {
        'book': book,
        'is_favorite': is_favorite,
        'reviews': reviews,
        'user_review': user_review,
        'form': form,
        'is_local': is_local,
        'similar_books': similar_books,
        'reading_progress': reading_progress,
    }
    return render(request, 'books/book_detail.html', context)
@login_required
def add_review(request, slug):
    book = get_object_or_404(Book, slug=slug)
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            Review.objects.update_or_create(
                book=book, user=request.user,
                defaults={
                    'rating': form.cleaned_data['rating'],
                    'comment': form.cleaned_data['comment']
                }
            )
            if request.headers.get('HX-Request'):
                reviews = book.reviews.all()
                response = render(request, 'books/partials/review_list_partial.html', {'reviews': reviews, 'book': book})
                response['HX-Trigger'] = 'reviewAdded'
                return response
            
            messages.success(request, "Votre avis a été enregistré.")
        else:
            if request.headers.get('HX-Request'):
                return render(request, 'books/partials/review_list_partial.html', {
                    'reviews': book.reviews.all(), 
                    'book': book,
                    'error': "Veuillez sélectionner une note et écrire un commentaire."
                })
            messages.error(request, "Veuillez corriger les erreurs dans votre avis.")
            
    return redirect('books:book_detail', slug=slug)

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def export_books(request):
    format_type = request.GET.get('format', 'excel')
    books = Book.objects.all().select_related('author', 'category').values(
        'title', 'author__name', 'category__name', 'isbn', 
        'publication_year', 'copies_available', 'is_available'
    )
    
    headers = ['Titre', 'Auteur', 'Catégorie', 'ISBN', 'Année', 'Stock', 'Disponible']
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="catalogue_estim_library.csv"'
        
        # Write UTF-8 BOM for Excel compatibility
        response.write(b'\xef\xbb\xbf')
        
        writer = csv.writer(response)
        writer.writerow(headers)
        for b in books:
            writer.writerow([
                b['title'], b['author__name'], b['category__name'], 
                b['isbn'], b['publication_year'], b['copies_available'],
                "Oui" if b['is_available'] else "Non"
            ])
        return response
    else:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="catalogue_estim_library.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogue"
        
        # Write headers
        ws.append(headers)
        
        # Write data
        for b in books:
            ws.append([
                b['title'], b['author__name'], b['category__name'], 
                b['isbn'], b['publication_year'], b['copies_available'],
                "Oui" if b['is_available'] else "Non"
            ])
            
        wb.save(response)
        return response

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def import_books(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            rows = []
            if file.name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(decoded_file)
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                header_row = [cell.value for cell in ws[1]]
                for row_data in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(header_row, row_data)))
            
            from .models import Author
            count = 0
            for row in rows:
                # Flexible column mapping
                title = row.get('Titre') or row.get('title')
                author_name = row.get('Auteur') or row.get('author')
                category_name = row.get('Catégorie') or row.get('category')
                isbn = str(row.get('ISBN') or row.get('isbn') or '')
                year = row.get('Année') or row.get('year') or 2024
                stock = row.get('Stock') or row.get('copies') or 1
                
                if title and author_name:
                    author, _ = Author.objects.get_or_create(name=author_name)
                    category = None
                    if category_name:
                        category, _ = Category.objects.get_or_create(name=category_name)
                    
                    Book.objects.get_or_create(
                        title=title,
                        author=author,
                        defaults={
                            'category': category,
                            'isbn': isbn,
                            'publication_year': int(year) if year else 2024,
                            'copies_available': int(stock) if stock else 1,
                            'is_available': True
                        }
                    )
                    count += 1
            
            messages.success(request, f"Importation réussie : {count} ouvrages ajoutés ou mis à jour.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {str(e)}")
            
    return redirect('books:manage_books')

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def fetch_book_info(request):
    isbn = request.GET.get('isbn')
    if not isbn:
        return JsonResponse({'error': 'ISBN manquant'}, status=400)
    
    isbn = isbn.replace('-', '').replace(' ', '')
    
    # --- STEP 1: Try Google Books (Fast, but has quotas) ---
    google_url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
    try:
        response = requests.get(google_url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if data.get('totalItems', 0) > 0:
                volume_info = data['items'][0]['volumeInfo']
                title = volume_info.get('title', '')
                authors = volume_info.get('authors', [])
                author_name = authors[0] if authors else "Auteur Inconnu"
                description = volume_info.get('description', '')
                pub_date = volume_info.get('publishedDate', '')
                pub_year = pub_date[:4] if pub_date else 2024
                
                from .models import Author, Category
                author, _ = Author.objects.get_or_create(name=author_name)
                
                # Category / Subject Extraction
                category_name_found = ""
                categories = volume_info.get('categories', [])
                if categories:
                    category_name_found = categories[0]
                
                category_id = None
                if category_name_found:
                    matched_cat = Category.objects.filter(name__iexact=category_name_found).first()
                    if matched_cat:
                        category_id = matched_cat.id

                return JsonResponse({
                    'success': True,
                    'source': 'Google',
                    'title': title,
                    'author_id': author.id,
                    'author_name': author.name,
                    'description': description,
                    'publication_year': pub_year,
                    'category_id': category_id,
                    'suggested_category': category_name_found
                })
    except Exception:
        pass # Silently fail and try fallback

    # --- STEP 2: Fallback to Open Library (No quota, but can be slower) ---
    ol_url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
    try:
        response = requests.get(ol_url, timeout=15)
        if response.status_code == 200:
            data = response.json()
            bib_key = f'ISBN:{isbn}'
            if bib_key in data:
                book_data = data[bib_key]
                title = book_data.get('title', '')
                authors = book_data.get('authors', [])
                author_name = authors[0].get('name') if authors else "Auteur Inconnu"
                
                description = book_data.get('notes', '')
                
                # Always try to get subjects regardless of description
                category_name_found = ""
                if 'subjects' in book_data:
                    subjects = [s.get('name') for s in book_data['subjects'][:5]]
                    category_name_found = subjects[0] if subjects else ""
                    if not description:
                        description = "Sujets : " + ", ".join(subjects)
                
                pub_date = book_data.get('publish_date', '')
                import re
                year_match = re.search(r'\d{4}', pub_date)
                pub_year = year_match.group(0) if year_match else 2024
                
                from .models import Author, Category
                author, _ = Author.objects.get_or_create(name=author_name)
                
                category_id = None
                if category_name_found:
                    matched_cat = Category.objects.filter(name__iexact=category_name_found).first()
                    if matched_cat:
                        category_id = matched_cat.id

                return JsonResponse({
                    'success': True,
                    'source': 'OpenLibrary',
                    'title': title,
                    'author_id': author.id,
                    'author_name': author.name,
                    'description': description,
                    'publication_year': pub_year,
                    'category_id': category_id,
                    'suggested_category': category_name_found
                })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f"Erreur de connexion (Google & OpenLibrary) : {str(e)}"})

    return JsonResponse({'success': False, 'error': 'Aucun livre trouvé pour cet ISBN sur aucune plateforme.'})

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def create_author_api(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            from .models import Author
            author, created = Author.objects.get_or_create(name=name)
            return JsonResponse({
                'success': True, 
                'id': author.id, 
                'name': author.name,
                'created': created
            })
    return JsonResponse({'success': False, 'error': 'Nom d\'auteur manquant ou méthode invalide.'}, status=400)

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def create_category_api(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            category, created = Category.objects.get_or_create(name=name)
            return JsonResponse({
                'success': True, 
                'id': category.id, 
                'name': category.name,
                'created': created
            })
    return JsonResponse({'success': False, 'error': 'Nom de catégorie manquant ou méthode invalide.'}, status=400)

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def manage_books(request):
    """
    Interface de gestion administrative des ouvrages.
    Optimisé avec select_related pour éviter les requêtes N+1 sur les auteurs et catégories.
    """
    # Base Queryset avec pré-chargement des relations
    books = Book.objects.all().select_related('author', 'category').order_by('-created_at')
    
    # Récupération des filtres depuis la requête GET
    query = request.GET.get('q')
    category_id = request.GET.get('category')
    status = request.GET.get('status')
    
    if query:
        books = books.filter(
            Q(title__icontains=query) | 
            Q(author__name__icontains=query) |
            Q(isbn__icontains=query)
        )
    
    if category_id:
        books = books.filter(category_id=category_id)
        
    # Logique de filtrage par état de stock ou popularité
    if status == 'available':
        books = books.filter(is_available=True)
    elif status == 'unavailable':
        books = books.filter(is_available=False)
    elif status == 'critical':
        # Stock critique : moins de 2 exemplaires physiques, sans version numérique
        books = books.filter(copies_available__lt=2, pdf_file='')
    elif status == 'popular':
        # Livres ayant plus de 5 réservations
        from django.db.models import Count
        books = books.annotate(res_count=Count('reservations')).filter(res_count__gt=5)

    # Calcul des statistiques pour les widgets du dashboard
    total_count = Book.objects.count()
    available_count = Book.objects.filter(is_available=True).count()
    availability_rate = (available_count / total_count * 100) if total_count > 0 else 0
    critical_count = Book.objects.filter(copies_available__lt=2, pdf_file='').count()
    total_categories = Category.objects.count()
    
    # Identification des IDs populaires pour l'affichage des badges
    from django.db.models import Count
    popular_ids = Book.objects.annotate(res_count=Count('reservations')).filter(res_count__gt=10).values_list('id', flat=True)

    # Mise en place de la pagination (20 livres par page)
    from django.core.paginator import Paginator
    paginator = Paginator(books, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'books': page_obj,
        'page_obj': page_obj,
        'categories': Category.objects.all(),
        'total_count': total_count,
        'availability_rate': round(availability_rate, 1),
        'critical_count': critical_count,
        'total_categories': total_categories,
        'popular_ids': popular_ids,
        'selected_category': category_id,
        'selected_status': status,
        'search_query': query,
    }
    
    # Support HTMX pour le filtrage en temps réel et le scroll infini
    if request.headers.get('HX-Request'):
        if request.headers.get('HX-Target') == 'books-table-container':
            return render(request, 'books/partials/manage_book_list_partial.html', context)
        return render(request, 'books/partials/manage_book_list_partial.html', {**context, 'infinite_scroll': True})
        
    return render(request, 'books/manage_books.html', context)

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def bulk_action_books(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        book_ids = request.POST.getlist('book_ids')
        
        if not book_ids:
            messages.warning(request, "Aucun livre sélectionné.")
            return redirect('books:manage_books')
            
        if action == 'delete':
            Book.objects.filter(id__in=book_ids).delete()
            messages.success(request, f"{len(book_ids)} ouvrages ont été supprimés.")
        elif action == 'online':
            Book.objects.filter(id__in=book_ids).update(is_available=True)
            messages.success(request, f"{len(book_ids)} ouvrages sont désormais en ligne.")
        elif action == 'offline':
            Book.objects.filter(id__in=book_ids).update(is_available=False)
            messages.success(request, f"{len(book_ids)} ouvrages ont été mis hors ligne.")
            
    return redirect('books:manage_books')

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def add_book(request):
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "L'ouvrage a été ajouté avec succès.")
            return redirect('books:manage_books')
    else:
        form = BookForm()
    return render(request, 'books/book_form.html', {'form': form, 'title': 'Ajouter un livre'})

@login_required
@user_passes_test(lambda u: u.role in ['admin', 'librarian'] or u.is_staff)
def edit_book(request, slug):
    book = get_object_or_404(Book, slug=slug)
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{book.title}' a été mis à jour.")
            return redirect('books:manage_books')
    else:
        form = BookForm(instance=book)
    return render(request, 'books/book_form.html', {'form': form, 'title': 'Modifier le livre', 'book': book})

@login_required
def read_book(request, slug):
    book = get_object_or_404(Book, slug=slug)
    if not book.pdf_file:
        return redirect('books:book_detail', slug=slug)
    
    progress, _ = ReadingProgress.objects.get_or_create(user=request.user, book=book)
    chapters = book.chapters.all()
    bookmarks = Bookmark.objects.filter(user=request.user, book=book)
    annotations = Annotation.objects.filter(user=request.user, book=book)
    
    context = {
        'book': book, 
        'progress': progress,
        'chapters': chapters,
        'bookmarks': bookmarks,
        'annotations': annotations,
    }
    return render(request, 'books/read_book.html', context)

@login_required
def add_bookmark(request, slug):
    if request.method == 'POST':
        book = get_object_or_404(Book, slug=slug)
        page = request.POST.get('page')
        label = request.POST.get('label', '')
        if page:
            Bookmark.objects.create(user=request.user, book=book, page_number=page, label=label)
            messages.success(request, f"Marque-page ajouté à la page {page}.")
    return redirect('books:read_book', slug=slug)

@login_required
def delete_bookmark(request, bookmark_id):
    bookmark = get_object_or_404(Bookmark, id=bookmark_id, user=request.user)
    slug = bookmark.book.slug
    bookmark.delete()
    return redirect('books:read_book', slug=slug)

@login_required
def add_annotation(request, slug):
    if request.method == 'POST':
        book = get_object_or_404(Book, slug=slug)
        page = request.POST.get('page')
        content = request.POST.get('content')
        if page and content:
            Annotation.objects.create(user=request.user, book=book, page_number=page, content=content)
            messages.success(request, "Note enregistrée.")
    return redirect('books:read_book', slug=slug)

@login_required
def delete_annotation(request, annotation_id):
    annotation = get_object_or_404(Annotation, id=annotation_id, user=request.user)
    slug = annotation.book.slug
    annotation.delete()
    return redirect('books:read_book', slug=slug)

@login_required
def update_reading_progress(request, slug):
    if request.method == 'POST':
        page = request.POST.get('page')
        if page:
            book = get_object_or_404(Book, slug=slug)
            progress, _ = ReadingProgress.objects.get_or_create(user=request.user, book=book)
            progress.last_page = page
            progress.save()
            
            if request.headers.get('HX-Request') or request.headers.get('Accept') == 'application/json':
                return JsonResponse({'status': 'success', 'page': page})
                
            return redirect('books:read_book', slug=slug)
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def toggle_favorite(request, slug):
    book = get_object_or_404(Book, slug=slug)
    favorite, created = Favorite.objects.get_or_create(user=request.user, book=book)
    
    if not created:
        favorite.delete()
        messages.info(request, f"'{book.title}' a été retiré de vos favoris.")
    else:
        messages.success(request, f"'{book.title}' a été ajouté à vos favoris !")
        
    return redirect(request.META.get('HTTP_REFERER', reverse('books:book_list')))

def catalog(request):
    """
    Vue principale du catalogue public.
    Implémente un filtrage intelligent par campus et département.
    """
    # Récupération des paramètres de filtrage depuis l'URL (GET)
    query = request.GET.get('q')
    category_id = request.GET.get('category')
    dept = request.GET.get('department')
    level = request.GET.get('level')
    campus_code = request.GET.get('campus')
    
    # Logique d'auto-filtrage au premier chargement pour les étudiants connectés
    # Si aucun filtre n'est spécifié, on utilise le campus et département de l'utilisateur
    is_first_load = not any([query, category_id, dept, level, campus_code])
    
    if is_first_load and request.user.is_authenticated:
        if request.user.campus:
            campus_code = request.user.campus.code
        if request.user.department:
            dept = request.user.department

    # Base Queryset avec pré-chargement des relations essentielles (Auteur, Catégorie)
    # On ne récupère que les livres marqués comme disponibles.
    books = Book.objects.filter(is_available=True).select_related('author', 'category')
    
    # Application des filtres de recherche textuelle
    if query:
        books = books.filter(
            Q(title__icontains=query) | 
            Q(author__name__icontains=query) |
            Q(description__icontains=query)
        )
    
    if category_id:
        books = books.filter(category_id=category_id)
        
    # Inclusion des ouvrages "Tout Public" (ceux sans département assigné)
    if dept:
        books = books.filter(Q(target_department=dept) | Q(target_department__isnull=True) | Q(target_department=''))
        
    if level:
        books = books.filter(target_level=level)

    # Filtrage par campus (Inclusion systématique des ressources globales 'all' ou sans campus)
    if campus_code:
        books = books.filter(
            Q(target_campuses__code=campus_code) | 
            Q(target_campuses__code='all') | 
            Q(target_campuses__isnull=True)
        ).distinct()

    # Tri par date de création (du plus récent au plus ancien)
    books = books.distinct().order_by('-created_at', 'id')

    # Données pour les listes déroulantes de filtres
    categories = Category.objects.all()
    campuses = Campus.objects.exclude(code='all') # On n'affiche pas 'all' dans les boutons
    
    # Pagination Infinite Scroll via HTMX (12 livres par lot)
    from django.core.paginator import Paginator
    paginator = Paginator(books, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Construction des tags pour afficher les filtres actuellement actifs
    active_filters = []
    selected_category_name = ""
    if query: active_filters.append({'type': 'q', 'label': f'Recherche: {query}'})
    if category_id: 
        cat = Category.objects.filter(id=category_id).first()
        if cat: 
            active_filters.append({'type': 'category', 'label': cat.name})
            selected_category_name = cat.name
    if dept:
        dept_label = dict(Book.DEPARTMENT_CHOICES).get(dept)
        active_filters.append({'type': 'department', 'label': dept_label})
    if campus_code:
        camp = Campus.objects.filter(code=campus_code).first()
        if camp: active_filters.append({'type': 'campus', 'label': camp.name})

    context = {
        'books': page_obj,
        'page_obj': page_obj,
        'categories': categories,
        'departments': Book.DEPARTMENT_CHOICES,
        'campuses': campuses,
        'search_query': query,
        'selected_category': category_id,
        'selected_category_name': selected_category_name,
        'selected_department': dept,
        'selected_campus': campus_code,
        'active_filters': active_filters,
    }
    
    # Réponses partielles HTMX (Scroll infini ou filtrage dynamique)
    if request.headers.get('HX-Request'):
        if request.headers.get('HX-Target') == 'catalog-container':
            return render(request, 'books/catalog.html', context)
        return render(request, 'books/partials/book_list_partial.html', context)
        
    return render(request, 'books/catalog.html', context)
